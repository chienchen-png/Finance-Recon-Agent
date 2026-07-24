"""
fix_executor.py - 本地修正执行器

接收修正方案 JSON 路径，在本地完成 Excel 修正、备份和验证。
AI 只读取执行摘要，不读取原始单元格值。
"""

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from backup_manager import create_single_snapshot


def _safe_text(value):
    if value is None:
        return ''
    return str(value).strip()


def _safe_number(value):
    if value is None or value == '':
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _resolve_project_root(path):
    for current in [path.parent, *path.parents]:
        if (current / '工作状态.json').exists() and (current / '结果输出').exists():
            return current
    return path.parents[3] if len(path.parents) > 3 else path.parent


def execute_fix_plan(plan_json_path, target_file, output_file, backup_dir,
                     operator_info=None, use_working_copy=False, quarter=None,
                     working_state_path=None):
    """
    执行本地修正，只返回安全摘要。

    V2.1 工作副本模式:
        use_working_copy=True 时，若 output_file 已存在则直接在其上修改，
        不再从 target_file 复制。第一次执行时（文件不存在）仍从 target_file 复制。
        每次写入前自动创建带季度标记的快照。
    """
    operator_info = operator_info or {}
    plan = json.loads(Path(plan_json_path).read_text(encoding='utf-8-sig'))
    fixes = plan.get('fixes', [])

    output_path_obj = Path(output_file)
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)

    # V2.1: 工作副本模式 — 仅首次从原数据复制
    if use_working_copy and output_path_obj.exists():
        # 工作副本已存在，直接在其上修改，不重新复制
        pass
    else:
        # 首次创建：从原数据复制
        shutil.copy2(target_file, output_file)

    # 首版原始快照（仅首次，幂等）
    backup_path = create_single_snapshot(target_file, backup_dir)

    # V2.1: 每季度修正前创建带季度标记的快照（工作副本模式下）
    quarterly_snapshot = None
    if use_working_copy and quarter and output_path_obj.exists():
        snap_name = f"snap_{quarter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        snap_path = Path(backup_dir) / snap_name
        shutil.copy2(output_file, snap_path)
        quarterly_snapshot = str(snap_path)

    # 文件锁检测：尝试在目标目录创建临时文件
    import tempfile as _tempfile
    try:
        _tmp = _tempfile.NamedTemporaryFile(dir=Path(output_file).parent, delete=False)
        _tmp_path = _tmp.name
        _tmp.close()
        Path(_tmp_path).unlink()
    except OSError:
        raise PermissionError(
            f"目标目录不可写，可能被其他程序（如 Excel）占用。请关闭 Excel 后重试。\n"
            f"文件: {output_file}"
        )

    wb = load_workbook(output_file)
    ws = wb.active

    executed = []
    skipped = []

    for item in fixes:
        target_row = item.get('target_row')
        target_col = item.get('target_col')
        new_value = item.get('new_value')
        if target_row is None or target_col is None:
            skipped.append({'contract': item.get('contract', ''), 'reason': '缺少定位信息'})
            continue

        if not isinstance(target_row, int) or target_row < 1:
            skipped.append({
                'contract': item.get('contract', ''),
                'reason': 'target_row 必须是 Excel 1-based 正整数',
            })
            continue

        if not isinstance(target_col, int) or target_col < 0:
            skipped.append({
                'contract': item.get('contract', ''),
                'reason': 'target_col 必须是 0-based 非负整数',
            })
            continue

        expected_contract = _safe_text(item.get('contract'))
        expected_person = _safe_text(item.get('person'))
        actual_contract = _safe_text(ws.cell(row=target_row, column=4).value)
        actual_person = _safe_text(ws.cell(row=target_row, column=3).value)
        strict = plan.get('strict_validation', True)
        
        if strict and expected_contract and expected_contract != actual_contract:
            skipped.append({
                'contract': expected_contract,
                'target_row': target_row,
                'reason': f'目标行合同不一致: expected={expected_contract}, actual={actual_contract}',
            })
            continue
        if strict and expected_person and expected_person != actual_person:
            skipped.append({
                'contract': expected_contract,
                'target_row': target_row,
                'reason': f'目标行销售人员不一致: expected={expected_person}, actual={actual_person}',
            })
            continue

        ws.cell(row=target_row, column=target_col + 1, value=new_value)
        executed.append({
            'contract': item.get('contract', ''),
            'person': item.get('person', ''),
            'level': item.get('level', ''),
            'target_row': target_row,
            'target_col': target_col,
            'new_value': new_value,
        })

    try:
        wb.save(output_file)
    except PermissionError:
        raise PermissionError(
            f"文件保存失败，{output_file} 可能被 Excel 占用。请关闭 Excel 后重试。"
        )
    wb.close()

    verify_wb = load_workbook(output_file, data_only=True)
    verify_ws = verify_wb.active
    verified = []
    for item in executed:
        cell_value = verify_ws.cell(row=item['target_row'], column=item['target_col'] + 1).value
        verified.append({
            'contract': item['contract'],
            'level': item['level'],
            'target_row': item['target_row'],
            'target_col': item['target_col'],
            'verified_value': _safe_number(cell_value),
        })
    verify_wb.close()

    # V2.0: 自动触发反向交叉校验
    reverse_vfy_summary = None
    rv_config = plan.get('reverse_verification')
    if rv_config and rv_config.get('enabled'):
        try:
            from verification_engine import run_reverse_verification
            rv_result = run_reverse_verification(
                quarters=rv_config['quarters'],
                target_file=output_file,
                target_sheet=rv_config.get('target_sheet', ws.title),
                target_field_mapping=rv_config.get('target_field_mapping', {}),
                knowledge_base_dir=rv_config.get('knowledge_base_dir', '知识库'),
                key_fields=rv_config.get('key_fields', ['contract', 'person']),
                amount_tolerance=rv_config.get('amount_tolerance', 0.01),
                exclude_contracts=rv_config.get('exclude_contracts'),
            )
            reverse_vfy_summary = rv_result.get('summary', {})
        except Exception as e:
            reverse_vfy_summary = {'error': str(e), 'status': 'verification_skipped'}

    summary = {
        'executed_count': len(executed),
        'skipped_count': len(skipped),
        'backup_file': backup_path,
        'quarterly_snapshot': quarterly_snapshot,
        'output_file': output_file,
        'generated_at': datetime.now().isoformat(),
        'operator_name': operator_info.get('operator_name', ''),
        'reviewer_name': operator_info.get('reviewer_name', ''),
        'operation_code': operator_info.get('operation_code', ''),
        'working_copy_mode': use_working_copy,
    }

    result = {
        'summary': summary,
        'executed': executed,
        'skipped': skipped,
        'verified': verified,
    }
    if reverse_vfy_summary:
        result['reverse_verification'] = reverse_vfy_summary

    output_path = Path(output_file)
    project_root = _resolve_project_root(output_path)
    result_dir = project_root / '结果输出' / '核对结果' / '_工作文件' / '执行摘要'
    result_dir.mkdir(parents=True, exist_ok=True)
    result_path = result_dir / f'{output_path.stem}.执行摘要.json'
    result['summary']['execution_summary_file'] = str(result_path)
    result_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')

    # V2.1: 自动更新工作状态中的 current_working_copy
    if use_working_copy and working_state_path and quarter:
        try:
            _update_working_state(working_state_path, output_file, target_file, quarter, quarterly_snapshot)
        except Exception:
            pass  # 状态更新失败不阻断主流程

    return result


def _update_working_state(working_state_path, output_file, target_file, quarter, snapshot):
    """更新 工作状态.json 中的 current_working_copy 字段"""
    ws_path = Path(working_state_path)
    if not ws_path.exists():
        return
    state = json.loads(ws_path.read_text(encoding='utf-8-sig'))
    wc = state.get('current_working_copy', {})
    now = datetime.now().isoformat()
    quarters = list(wc.get('quarters_applied', []))
    if quarter not in quarters:
        quarters.append(quarter)
    snap_chain = list(wc.get('snapshot_chain', []))
    if snapshot:
        snap_chain.append({'quarter': quarter, 'snapshot': snapshot})
    state['current_working_copy'] = {
        'path': output_file,
        'created_from': wc.get('created_from') or target_file,
        'created_at': wc.get('created_at') or now,
        'last_modified_at': now,
        'quarters_applied': quarters,
        'snapshot_chain': snap_chain,
    }
    state['last_updated'] = now
    ws_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding='utf-8-sig')


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='执行本地安全修正（V2.1 工作副本模式）')
    parser.add_argument('--config', required=True, help='修正配置 JSON 路径')
    args = parser.parse_args()

    config = json.loads(Path(args.config).read_text(encoding='utf-8-sig'))
    # V2.1: 配置 JSON 可包含 use_working_copy / quarter / working_state_path
    print(json.dumps(execute_fix_plan(**config), ensure_ascii=False, indent=2))
