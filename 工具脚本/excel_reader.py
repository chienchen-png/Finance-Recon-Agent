"""
excel_reader.py — 统一 Excel 读取器
兼容 .xls / .xlsx 格式，提供一致的读取接口。

被以下 Skill 调用: 01-表格扫描与结构化

所有路径均为相对于项目根目录的相对路径，由调用方拼接基础路径后传入。
"""

from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


def read_excel(file_path: str, sheet_name: str = None):
    """
    统一读取 Excel 文件，自动识别 .xls / .xlsx。

    Args:
        file_path: Excel 文件路径
        sheet_name: 指定 sheet 名，为 None 时返回所有 sheet

    Returns:
        dict: {sheet_name: [[row_data], ...], ...}
    """
    ext = Path(file_path).suffix.lower()

    if ext == '.xlsx':
        return _read_xlsx(file_path, sheet_name)
    elif ext == '.xls':
        return _read_xls(file_path, sheet_name)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")


def read_schema_only(file_path: str, header_row: int = 0, sheet_name: str = None,
                     data_start_row: int = None, preview_rows: int = 5) -> dict:
    """
    仅读取 Excel 结构元数据，不返回任何数据行或单元格明细值。

    V2.0 新增: 自适应导入支持 — 自动探测数据起始行、预览前 N 行样例值。

    Args:
        file_path: Excel 文件路径
        header_row: 表头所在行号，0-based
        sheet_name: 指定 sheet 名，为 None 时扫描所有 sheet
        data_start_row: 数据起始行号（0-based），None 时自动探测
        preview_rows: 预览行数，默认 5

    Returns:
        dict: 文件名、sheet 名、行列数、字段名、数据起始行建议、字段样例
    """
    ext = Path(file_path).suffix.lower()

    if ext == '.xlsx':
        result = _read_xlsx_schema(file_path, header_row, sheet_name, data_start_row, preview_rows)
    elif ext == '.xls':
        result = _read_xls_schema(file_path, header_row, sheet_name, data_start_row, preview_rows)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")

    return result


def _read_xlsx(file_path: str, sheet_name: str = None) -> dict:
    """读取 .xlsx 文件"""
    if openpyxl is None:
        raise ImportError("请先安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames

    result = {}
    for sn in sheets_to_read:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        result[sn] = rows

    wb.close()
    return result


def _read_xlsx_schema(file_path: str, header_row: int = 0, sheet_name: str = None,
                      data_start_row: int = None, preview_rows: int = 5) -> dict:
    """读取 .xlsx 文件结构，不返回数据行。V2.0 增强：自适应数据始行探测 + 字段样例返回。"""
    if openpyxl is None:
        raise ImportError("请先安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    sheets_to_read = [sheet_name] if sheet_name else wb.sheetnames

    result = {'file_name': Path(file_path).name, 'sheets': {}}
    for sn in sheets_to_read:
        ws = wb[sn]
        fields = []
        if 0 <= header_row < ws.max_row:
            header_cells = next(ws.iter_rows(
                min_row=header_row + 1,
                max_row=header_row + 1,
                values_only=True
            ))
            fields = [
                {'name': str(value).strip() if value is not None else '', 'col_index': index}
                for index, value in enumerate(header_cells)
            ]
        # 自动探测数据起始行
        detected_dsr = data_start_row
        field_samples = []
        suggestion_reason = '用户指定'
        if detected_dsr is None:
            detected_dsr = _detect_data_start_xlsx(ws, header_row)
            suggestion_reason = '自动探测'
        # 预览前 N 行
        field_names = [f['name'] for f in fields]
        if detected_dsr is not None and preview_rows > 0:
            for r_offset in range(preview_rows):
                pr = detected_dsr + r_offset + 1  # 1-based
                if pr > ws.max_row:
                    break
                row_vals = next(ws.iter_rows(min_row=pr, max_row=pr, values_only=True))
                sample = {}
                for fi, fv in enumerate(field_names):
                    if fi < len(row_vals):
                        sample[fv] = str(row_vals[fi])[:80] if row_vals[fi] is not None else ''
                field_samples.append({'row': pr, 'sample': sample})
        result['sheets'][sn] = {
            'row_count': ws.max_row,
            'col_count': ws.max_column,
            'header_row': header_row,
            'fields': fields,
            'field_names': field_names,
            'data_start_row': detected_dsr,
            'data_start_row_source': suggestion_reason,
            'field_samples': field_samples,
            'has_merged_cells': bool(ws.merged_cells.ranges),
        }

    wb.close()
    return result


def _detect_data_start_xlsx(ws, header_row: int) -> int:
    """自动探测数据起始行（跳过表头 + 空行 + 全合行标记行）"""
    candidate = header_row + 1
    max_check = min(candidate + 20, ws.max_row or candidate + 20)
    for r in range(candidate, max_check + 1):
        row_cells = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        all_str = ' '.join(str(c).strip() for c in row_cells if c is not None)
        if any(kw in all_str for kw in ['合  计', '小  计', '总计', '合计', '小计']):
            continue  # 跳过全合标记行
        non_empty = sum(1 for c in row_cells if c is not None and str(c).strip())
        if non_empty >= 2:
            # 找到第一个≥2非空列的行作为数据始行（1-based Excel行号→0-based索引）
            return r - 1
    return candidate


def _read_xls(file_path: str, sheet_name: str = None) -> dict:
    """读取 .xls 文件"""
    if xlrd is None:
        raise ImportError("请先安装 xlrd: pip install xlrd")

    wb = xlrd.open_workbook(file_path)
    sheets_to_read = [sheet_name] if sheet_name else wb.sheet_names()

    result = {}
    for sn in sheets_to_read:
        ws = wb.sheet_by_name(sn)
        rows = []
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        result[sn] = rows

    return result


def _read_xls_schema(file_path: str, header_row: int = 0, sheet_name: str = None,
                     data_start_row: int = None, preview_rows: int = 5) -> dict:
    """读取 .xls 文件结构，不返回数据行。V2.0 增强：自适应数据始行探测 + 字段样例返回。"""
    if xlrd is None:
        raise ImportError("请先安装 xlrd: pip install xlrd")

    wb = xlrd.open_workbook(file_path)
    sheets_to_read = [sheet_name] if sheet_name else wb.sheet_names()

    result = {'file_name': Path(file_path).name, 'sheets': {}}
    for sn in sheets_to_read:
        ws = wb.sheet_by_name(sn)
        fields = []
        if 0 <= header_row < ws.nrows:
            fields = [
                {'name': str(ws.cell_value(header_row, col)).strip(), 'col_index': col}
                for col in range(ws.ncols)
            ]
        field_names = [f['name'] for f in fields]
        # 自动探测数据起始行
        detected_dsr = data_start_row
        suggestion_reason = '用户指定'
        if detected_dsr is None:
            detected_dsr = _detect_data_start_xls(ws, header_row)
            suggestion_reason = '自动探测'
        # 预览前 N 行
        field_samples = []
        if detected_dsr is not None and preview_rows > 0:
            for r_offset in range(preview_rows):
                pr = detected_dsr + r_offset
                if pr >= ws.nrows:
                    break
                sample = {}
                for fi, fv in enumerate(field_names):
                    if fi < ws.ncols:
                        val = ws.cell_value(pr, fi)
                        sample[fv] = str(val)[:80] if val != '' else ''
                field_samples.append({'row': pr, 'sample': sample})
        result['sheets'][sn] = {
            'row_count': ws.nrows,
            'col_count': ws.ncols,
            'header_row': header_row,
            'fields': fields,
            'field_names': field_names,
            'data_start_row': detected_dsr,
            'data_start_row_source': suggestion_reason,
            'field_samples': field_samples,
            'has_merged_cells': False,
        }

    return result


def _detect_data_start_xls(ws, header_row: int) -> int:
    """自动探测 .xls 数据起始行（0-based）"""
    candidate = header_row + 1
    max_check = min(candidate + 20, ws.nrows - 1)
    for r in range(candidate, max_check + 1):
        all_str = ' '.join(
            str(ws.cell_value(r, c)).strip()
            for c in range(min(ws.ncols, 10))
        )
        if any(kw in all_str for kw in ['合  计', '小  计', '总计', '合计', '小计']):
            continue
        non_empty = sum(
            1 for c in range(ws.ncols)
            if ws.cell_value(r, c) != '' and str(ws.cell_value(r, c)).strip()
        )
        if non_empty >= 2:
            return r
    return candidate


def get_sheet_names(file_path: str) -> list:
    """获取 Excel 中所有 sheet 名"""
    return list(read_excel(file_path).keys())


def list_excel_files(directory: str) -> list:
    """列出目录下所有 Excel 文件"""
    patterns = ['*.xlsx', '*.xls', '*.xlsm']
    files = []
    for pattern in patterns:
        files.extend(Path(directory).glob(pattern))
    return [str(f) for f in files]
